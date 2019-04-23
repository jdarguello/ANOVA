#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from outliers import smirnov_grubbs as grubbs
import pandas as pd
import numpy as np
import scipy.stats
import math
import matplotlib.pyplot as plt
import openpyxl as op
from PIL import Image

class TxtData():
    """
    Obtiene datos generales
    """
    
    def txtinput(self):
        with open("Generalidades.txt", "r") as file:
                frases = file.read().split("\n")
                data = {}
                for frase in frases:
                        if len(frase) > 0 and frase[:1] != "#":
                            contenido = frase.split(' ')
                            try:
                                data[contenido[0]] = float(contenido[2])
                            except:
                                if contenido[2] == "True":
                                    data[contenido[0]] = True
                                elif contenido[2] == "False":
                                    data[contenido[0]] = False
                                else:
                                    data[contenido[0]] = contenido[2]
                return data

class ExcelIO():
    """
    Obtiene los datos y envía resultados en formato .xlsx
    """

    def Input(self, file):
            #Importación del archivo
            xl_file = pd.read_excel(file)
            Data = pd.DataFrame(xl_file)

            #Procesamiento de NaN
            #Data = Data.replace(np.nan, 0)

            #Datos relevantes
            Relevant_Data = Data.iloc[:, 1:]
            return Relevant_Data

    def colnum_string(self, n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def Output(self, res, file, txtDat):
            wb = op.load_workbook(file)
            indeseables = ['GAVG', 'Q', 'Qc', 'Qw', 'QA', 't', 'n',
                'S2A', 'S2w', 'S2', 'f', 'fumbral']
            #----CREACIÓN DE HOJAS----
            #Cambio de nombre 'Hoja 1'
            if 'Hoja1' in wb.sheetnames:
                hoja1 = wb.get_sheet_by_name('Hoja1')
                hoja1.title = 'Datos'
                wb.save(file)
            elif 'Sheet1' in wb.sheetnames:
                sheet1 = wb.get_sheet_by_name('Sheet1')
                sheet1.title = 'Data'
                wb.save(file)
            #-----NormalDist-----
            #Creación hoja NormalDist, si no existe
            if 'NormalDist' in wb.sheetnames:
                std = wb.get_sheet_by_name('NormalDist')
                wb.remove_sheet(std)
            ND =  wb.create_sheet('NormalDist')
            wb.save(file)
            #Escritura de resultados
            last = 1
            for key, values in res.items():
                validate = True
                for ind in indeseables:
                    if key == ind:
                        validate = False
                if validate:
                    #print(key)
                    original = res[key]['Datos originales']
                    for i in range(original['cant_datos']):
                        if i == 0:
                            #Resize
                            imgdir = direc + key + '.png'
                            ancho = 575
                            alto = 345
                            size = (ancho, alto)
                            img = Image.open(imgdir)
                            img.thumbnail(size, Image.ANTIALIAS)
                            img.save(imgdir)

                            img = op.drawing.image.Image(imgdir)
                            ND.add_image(img, 'F' + str(last))

                            ND['A' + str(last)] = key
                            last += 1

                            ND['A' + str(last)] = 'Datos'
                            ND['B' + str(last)] = 'Orden'
                            ND['C' + str(last)] = 'Fracción'
                            ND['D' + str(last)] = 'Z'
                        else:
                            ND['A' + str(last)] = original['data'][i]
                            ND['B' + str(last)] = original['Orden'][i]
                            ND['C' + str(last)] = original['Fracción'][i]
                            ND['D' + str(last)] = original['Z'][i]
                        last += 1
                    last += 10
            wb.save(file)
            #----DATOS DEFINITIVOS----
            #Creación hoja 'Datos definitivos'
            if 'Datos definitivos' in wb.sheetnames:
                std = wb.get_sheet_by_name('Datos definitivos')
                wb.remove_sheet(std)
            DD = wb.create_sheet('Datos definitivos')
            wb.save(file)
            #Poner valores
            print(res)
            cant_datos = 0
            cont = 2
            last = 1
            for key, value in res.items():
                data = False
                try:
                    data = res[key]['Datos definitivos']
                except:
                    pass
                if data:
                    if data['cant_datos'] > cant_datos:
                        cant_datos = data['cant_datos']
                    letter = self.colnum_string(cont)
                    DD[letter + '1'] = key
                    for i in range(data['cant_datos']):
                        DD[letter + str(i+2)] = data['data'][i]
                    cont += 1
            DD['A1'] = "Réplicas"
            for i in range(cant_datos):
                DD['A' + str(i+2)] = i + 1 
            wb.save(file)
            #-----Qs-----
            #Creación hoja 'Qs'
            if 'Qs' in wb.sheetnames:
                std = wb.get_sheet_by_name('Qs')
                wb.remove_sheet(std)
            Qs = wb.create_sheet('Qs')
            wb.save(file)
            Qs['A1'] = 
                    
                    

class ANOVA1(ExcelIO, TxtData):
    def __init__(self, **kwargs):
            #Lectura de Generalidades.txt
            txt_Dat = self.txtinput()

            #Datos de interés
            Data = self.Input(kwargs['file'])
            Encabezados = list(Data)
            #Datos estadísticos generales
            General = self.EG(Data, Encabezados)

            #Eliminación de datos anómalos
            for columna in Encabezados:
                if General[columna]['Datos originales']['cant_datos'] <= 15 or \
                        txt_Dat['Grubbs']:
                        self.GrubbsAnomaly(General, txt_Dat['alpha'], columna)
                else:
                        self.NormalDist(General[columna], kwargs['directory'], columna)
                if txt_Dat['NormalDist']:
                        self.NormalDist(General[columna], kwargs['directory'], columna)
            #Promedio global
            General['GAVG'] = self.Average(General, Encabezados)

            #Cálculo de las Q's
            General['Qw'] = 0
            General['QA'] = 0
            General['Q'] = 0   
            for columna in Encabezados:
                self.Qs(General[columna], General)
            General['Qc'] = General['QA']+General['Qw']    #Q comparativo

            #t y n
            t = 0
            n = 0
            for columna in Encabezados:
                t += 1
                n += General[columna]['Datos definitivos']['cant_datos']
            General['t'] = t
            General['n'] = n

            #S2
            General['S2A'] = General['QA']/(t-1)
            General['S2w'] = General['Qw']/(n-t)
            General['S2'] = General['Q']/(n-1)
            
            #f y fumbral
            General['f'] = General['S2A']/General['S2w']
            General['fumbral'] = scipy.stats.f.isf(txt_Dat['precision_f'], t-1, n-t)
            #print(General)

            #Resultados
            self.Output(General, kwargs['file'], txt_Dat)

    def Qs(self, Data, General):
            #Qw
            Data['Qws'] = np.zeros(len(Data['Datos definitivos']['data']))
            for i in range(len(Data['Datos definitivos']['data'])):
                Data['Qws'][i] = Data['Datos definitivos']['data'][i]-Data['Datos definitivos']['promedio']
                General['Qw'] += Data['Qws'][i]**2
            #Q
            Data['Qs'] = np.zeros(len(Data['Datos definitivos']['data']))
            for i in range(len(Data['Datos definitivos']['data'])):
                Data['Qs'][i] = Data['Datos definitivos']['data'][i]-General['GAVG']
                General['Q'] += Data['Qs'][i]**2
            #QA
            General['QA'] += Data['Datos definitivos']['cant_datos']*(Data['dif_prom']**2)
            
    
    def Average(self, Data, Encabezados):
            #Promedio ponderado general     
            suma = 0
            num_datos = 0
            for columna in Encabezados:
                for dato in Data[columna]['Datos definitivos']['data']:
                        suma += dato
                num_datos += Data[columna]['Datos definitivos']['cant_datos']
            promedio = suma/num_datos
            #Diferencia entre promedio local y general
            for columna in Encabezados:
                Data[columna]['dif_prom'] = Data[columna]['Datos definitivos']['promedio']-promedio
            return promedio
        

    def NormalDist(self, Data, directory, columna):
            Data['Datos originales']['Orden'] = np.searchsorted(
                np.sort(Data['Datos originales']['data']),
                Data['Datos originales']['data'], side='left')
            #Creación de formatos de datos "Fracción", "Z", Log(P) y %
            Data['Datos originales']['Fracción'] = np.zeros(
                Data['Datos originales']['cant_datos'])
            Data['Datos originales']['Z'] = np.zeros(
                Data['Datos originales']['cant_datos'])
            Data['Datos originales']['logP'] = np.zeros(
                Data['Datos originales']['cant_datos'])
            Data['Datos originales']['Porcentaje'] = np.zeros(
                Data['Datos originales']['cant_datos'])
            #Llenado de la información
            for i in range(Data['Datos originales']['cant_datos']):
                Data['Datos originales']['Orden'][i] += 1
                Data['Datos originales']['Fracción'][i] = \
                (Data['Datos originales']['Orden'][i]-0.5)/Data\
                ['Datos originales']['cant_datos']
                Data['Datos originales']['Z'][i] = np.random.normal(
                Data['Datos originales']['Fracción'][i])
                Data['Datos originales']['logP'][i] = math.log10(
                100*Data['Datos originales']['Fracción'][i])
                Data['Datos originales']['Porcentaje'][i] = 100*\
                Data['Datos originales']['Fracción'][i]
            #GRÁFICAS
            #Prueba de distribución normal
            x = Data['Datos originales']['data']
            y = Data['Datos originales']['Fracción']
            fig, ax = plt.subplots()
            ax.plot(x,y,'.')
            #Línea de tendencia
            z = np.polyfit(x,y,1)
            p = np.poly1d(z)
            ax.plot(x,p(x), "r-")
            #Generalidades del gráfico
            plt.title("Prueba de distirbución normal - " + columna)
            plt.xlabel("Datos")
            plt.ylabel("Z")
            #plt.show()
            #Guardar gráfica
            plt.savefig(directory + columna + '.png')
            #print(Data)

    def GrubbsAnomaly(self, Data, alpha, columna):
            col = {}
            col['data'] = grubbs.test(Data[columna]\
                ['Datos originales']['data'], alpha=alpha)
            #Conteo de datos
            col['cant_datos'] = len(col['data'])
            col['promedio'] = np.mean(col['data'])
            col['std'] = np.std(col['data'])

            Data[columna]['Datos definitivos'] = col
    
    def EG(self, Data, Encabezados):
            #Conversión de datos a matriz
            Matriz = Data.values

            #Organización de datos en diccionario
            Datos = {}
            con = 0
            for columna in Encabezados:
                col = {}
                #Conteo de datos por columna
                contador = 0
                for i in range(Matriz.shape[0]):
                        if np.isnan(Matriz[i][con]):
                            pass
                        else:
                            contador += 1
                        dat = np.zeros(contador)
        
                #Reemplazar valores por reales en cada vector columna
                for i in range(contador):
                        dat[i] = Matriz[i][con]
        
                col['data'] = dat
                col['cant_datos'] = contador
                col['promedio'] = np.mean(dat)
                col['std'] = np.std(dat)

                original = {}
                original['Datos originales'] = col

                Datos[columna] = original
                con += 1
            return Datos

if __name__:
    direc = 'Datos/Ejemplo1/'
    file = direc + 'Ejemplo1.xlsx'
    ANOVA1(file=file, alpha=0.05, directory=direc)
